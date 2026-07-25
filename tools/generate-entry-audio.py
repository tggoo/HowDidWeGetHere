"""Generate legacy narration plaintext, MP3 files and upload ZIPs for workbook-derived entries."""

import argparse
import re
import subprocess
import sys
import unicodedata
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from markdown_tts import markdown_to_tts


SUPPORTED_SHEETS = ("Master Timeline", "Mythology Index")


def create_slug(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    chars: list[str] = []
    for char in normalized:
        if unicodedata.category(char) == "Mn":
            continue
        chars.append(char.lower() if char.isalnum() else "-")
    slug = re.sub("-+", "-", "".join(chars)).strip("-")
    return slug or "entry"


def read_rows(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    for sheet_name in SUPPORTED_SHEETS:
        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
        index = {header: offset for offset, header in enumerate(headers)}

        for row in iterator:
            if not any(value is not None and str(value).strip() for value in row):
                continue

            def value(header: str) -> str:
                offset = index.get(header)
                if offset is None or offset >= len(row) or row[offset] is None:
                    return ""
                return str(row[offset]).strip()

            if sheet_name == "Master Timeline":
                title = value("Event / development")
                summary = value("Why it matters")
                date_label = value("Approx. date")
                category = value("Category")
                region = value("Region")
            else:
                title = value("Figure / creature")
                summary = value("What it represents / famous story")
                date_label = value("Probable tradition age")
                category = value("Type")
                region = value("Tradition")

            if not title:
                continue

            rows.append(
                {
                    "slug": create_slug(title),
                    "title": title,
                    "summary": summary,
                    "date_label": date_label,
                    "category": category,
                    "region": region,
                    "sheet": sheet_name,
                }
            )
    return rows


def narration_text(row: dict[str, str], language: str) -> str:
    parts = [row["title"]]
    if row["date_label"]:
        parts.append(f"Approximate date: {row['date_label']}.")
    if row["region"]:
        parts.append(f"Region or tradition: {row['region']}.")
    if row["category"]:
        parts.append(f"Category: {row['category']}.")
    if row["summary"]:
        parts.append(row["summary"])
    return markdown_to_tts("\n\n".join(parts), language)


def run_tts(tts_root: Path, text_path: Path, mp3_path: Path, lang: str) -> None:
    python_path = tts_root / ".venv" / "Scripts" / "python.exe"
    if not python_path.exists():
        python_path = Path(sys.executable)
    script_path = tts_root / "text_to_speech.py"
    subprocess.run(
        [
            str(python_path),
            str(script_path),
            str(text_path),
            "--lang",
            lang,
            "--output",
            str(mp3_path),
        ],
        check=True,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate entry narration text, mp3 files and a bulk-upload zip.")
    parser.add_argument("--workbook", default="world_history_mythology_timeline_to_2026.xlsx")
    parser.add_argument("--tts-root", default=r"C:\Users\danie\Documents\Repos\TTS")
    parser.add_argument("--lang", default="en", choices=("en", "cs", "es"))
    parser.add_argument("--out", default="generated/audio")
    parser.add_argument("--generate", action="store_true", help="Call the TTS project and create mp3 files.")
    parser.add_argument("--limit", type=int, default=0, help="Limit processed rows for a test run.")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    tts_root = Path(args.tts_root)
    out_root = Path(args.out)
    out_root.mkdir(parents=True, exist_ok=True)

    rows = read_rows(workbook_path)
    if args.limit > 0:
        rows = rows[: args.limit]

    mp3_paths: list[Path] = []
    zip_entries: list[tuple[Path, str]] = []
    for row in rows:
        entry_root = out_root / row["slug"] / args.lang
        entry_root.mkdir(parents=True, exist_ok=True)
        text_path = entry_root / "narration.txt"
        mp3_path = entry_root / "narration.mp3"
        text_path.write_text(narration_text(row, args.lang), encoding="utf-8")
        if args.generate and (args.overwrite or not mp3_path.exists() or mp3_path.stat().st_size == 0):
            run_tts(tts_root, text_path, mp3_path, args.lang)
        if mp3_path.exists() and mp3_path.stat().st_size > 0:
            mp3_paths.append(mp3_path)
            zip_entries.append((mp3_path, f"{row['slug']}.{args.lang}.mp3"))

    zip_path = out_root / f"entry-audio-{args.lang}.zip"
    if mp3_paths:
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for mp3_path, archive_name in zip_entries:
                archive.write(mp3_path, arcname=archive_name)

    print(f"Wrote {len(rows)} text files under {out_root}.")
    if args.generate:
        print(f"Generated/found {len(mp3_paths)} mp3 files under {out_root}.")
    if mp3_paths:
        print(f"Wrote bulk upload zip: {zip_path}")


if __name__ == "__main__":
    main()
