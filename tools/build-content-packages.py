import argparse
import json
import re
import shutil
import unicodedata
import zipfile
from pathlib import Path

from openpyxl import load_workbook


SUPPORTED_SHEETS = {
    "Master Timeline": "master-timeline",
    "Mythology Index": "mythology",
}

LANGUAGE_ALIASES = {
    "en": "en",
    "cs": "cs",
    "cz": "cs",
    "es": "es",
    "sp": "es",
}

ERA_RANGES = {
    "prehistory": (-3000000, -3000),
    "neolithic": (-10000, -3300),
    "ancient": (-3300, 500),
    "late-antiquity": (250, 750),
    "middle-ages": (500, 1500),
    "early-modern": (1500, 1800),
    "industrial-age": (1760, 1914),
    "modern": (1800, 1945),
    "contemporary": (1945, 2026),
}

CS_LABELS = {
    "africa": "Afrika",
    "agriculture": "Zemědělství",
    "ambiguous god": "Nejednoznačné božstvo",
    "americas": "Ameriky",
    "america": "Ameriky",
    "anatolia": "Anatolie",
    "ancient": "Starověk",
    "andes": "Andy",
    "antarctica": "Antarktida",
    "arabia": "Arábie",
    "architecture": "Architektura",
    "asia": "Asie",
    "atlantic": "Atlantik",
    "atlantic ocean": "Atlantský oceán",
    "auspicious chimera": "Příznivá chiméra",
    "australia": "Austrálie",
    "britain": "Británie",
    "cat": "Kočka",
    "cattle god": "Dobytčí bůh",
    "celtic": "Keltská",
    "central and western europe": "Střední a západní Evropa",
    "chaos serpent": "Had chaosu",
    "chief god": "Hlavní bůh",
    "china": "Čína",
    "chinese": "Čínská",
    "civilization": "Civilizace",
    "conquest": "Dobývání",
    "contemporary": "Současnost",
    "creator goddess": "Stvořitelská bohyně",
    "creator kami": "Stvořitelské kami",
    "culture": "Kultura",
    "culture hero": "Kulturní hrdina",
    "deity": "Božstvo",
    "demon": "Démon",
    "desert, storm and disorder god": "Bůh pouště, bouře a nepořádku",
    "early modern": "Raný novověk",
    "earth goddess": "Bohyně země",
    "eastern mediterranean": "Východní Středomoří",
    "economy": "Ekonomika",
    "egypt": "Egypt",
    "egyptian": "Egyptská",
    "eight-headed serpent": "Osmihlavý had",
    "empire": "Říše",
    "end and renewal of the world": "Konec a obnova světa",
    "england": "Anglie",
    "environment": "Životní prostředí",
    "eurasia": "Eurasie",
    "europe": "Evropa",
    "exploration": "Objevování",
    "female death omen spirit": "Ženský duch věštící smrt",
    "fire deity": "Ohnivé božstvo",
    "folklore witch": "Folklorní čarodějnice",
    "forest spirit": "Lesní duch",
    "fox spirit": "Liščí duch",
    "france": "Francie",
    "funerary and embalming god": "Bůh pohřbívání a balzamování",
    "global": "Globální",
    "global oceans": "Světové oceány",
    "god": "Bůh",
    "god of death and rebirth": "Bůh smrti a znovuzrození",
    "goddess": "Bohyně",
    "goddess of magic and motherhood": "Bohyně magie a mateřství",
    "gorgon monster": "Gorgonská příšera",
    "greece": "Řecko",
    "greek": "Řecká",
    "hero": "Hrdina",
    "heroic king": "Hrdinský král",
    "high god": "Nejvyšší bůh",
    "historical event": "Historická událost",
    "horned god": "Rohatý bůh",
    "house spirit": "Domácí duch",
    "human history": "Dějiny lidstva",
    "human-horse hybrid": "Kříženec člověka a koně",
    "hunt and wilderness goddess": "Bohyně lovu a divočiny",
    "hybrid guardian": "Hybridní strážce",
    "hybrid monster": "Hybridní příšera",
    "iceland": "Island",
    "india": "Indie",
    "industrial age": "Průmyslový věk",
    "industrialization": "Industrializace",
    "infrastructure": "Infrastruktura",
    "international relations": "Mezinárodní vztahy",
    "invention": "Vynález",
    "ireland": "Irsko",
    "italy": "Itálie",
    "japan": "Japonsko",
    "japanese": "Japonská",
    "late antiquity": "Pozdní antika",
    "levant": "Levant",
    "lioness goddess": "Lví bohyně",
    "literature": "Literatura",
    "love goddess": "Bohyně lásky",
    "marriage and queenship goddess": "Bohyně manželství a královské moci",
    "medicine": "Medicína",
    "mediterranean": "Středomoří",
    "mesopotamia": "Mezopotámie",
    "mexico": "Mexiko",
    "middle ages": "Středověk",
    "middle east": "Blízký východ",
    "modern": "Moderní doba",
    "monkey king": "Opičí král",
    "monstrous wolf": "Obludný vlk",
    "moon": "Měsíc",
    "mountain spirit": "Horský duch",
    "mythic animal": "Mýtické zvíře",
    "mythic fox": "Mýtická liška",
    "mythology": "Mytologie",
    "nature spirit": "Přírodní duch",
    "neolithic": "Neolit",
    "nine-tailed fox": "Devítiocasá liška",
    "norse": "Severská",
    "north africa": "Severní Afrika",
    "north america": "Severní Amerika",
    "northern europe": "Severní Evropa",
    "ogre": "Zlobr",
    "one-eyed giants": "Jednoocí obři",
    "pacific": "Pacifik",
    "pacific ocean": "Tichý oceán",
    "pandemic": "Pandemie",
    "persia": "Persie",
    "philosophy": "Filozofie",
    "policy": "Veřejná politika",
    "political history": "Politické dějiny",
    "politics": "Politika",
    "prehistory": "Pravěk",
    "present day": "Současnost",
    "primordial state": "Prvotní stav",
    "prophecy, music and plague god": "Bůh proroctví, hudby a moru",
    "reborn bird": "Znovuzrozený pták",
    "religion": "Náboženství",
    "revolution": "Revoluce",
    "roman empire": "Římská říše",
    "roman judea": "Římská Judea",
    "rome": "Řím",
    "russia": "Rusko",
    "sacred bird": "Posvátný pták",
    "science": "Věda",
    "sea and earthquake god": "Bůh moře a zemětřesení",
    "sea kami": "Mořské kami",
    "sky": "Nebe",
    "sky and kingship god": "Bůh nebe a královské moci",
    "sky and thunder god": "Bůh nebe a hromu",
    "sky god": "Bůh nebe",
    "slavic": "Slovanská",
    "slavic europe": "Slovanská Evropa",
    "south asia": "Jižní Asie",
    "southwest asia": "Jihozápadní Asie",
    "southwest asia and other regions": "Jihozápadní Asie a další regiony",
    "sovereignty goddess": "Bohyně svrchovanosti",
    "space": "Vesmír",
    "storm": "Bouře",
    "sun god": "Bůh slunce",
    "sun goddess": "Bohyně slunce",
    "supernatural being": "Nadpřirozená bytost",
    "supernatural warrior women": "Nadpřirozené bojovnice",
    "technology": "Technologie",
    "terrorism": "Terorismus",
    "thunder god": "Bůh hromu",
    "titan ruler": "Vládce Titánů",
    "traditional history": "Tradiční dějiny",
    "transport": "Doprava",
    "trickster": "Šibal",
    "ukraine": "Ukrajina",
    "underworld": "Podsvětí",
    "underworld god": "Bůh podsvětí",
    "underworld queen": "Královna podsvětí",
    "united states": "Spojené státy",
    "vegetation goddess": "Bohyně vegetace",
    "wales": "Wales",
    "war": "Válka",
    "war and law god": "Bůh války a práva",
    "water": "Voda",
    "water horse spirit": "Vodní koňský duch",
    "western europe": "Západní Evropa",
    "wisdom and war goddess": "Bohyně moudrosti a války",
    "world serpent": "Světový had",
    "world tree": "Světový strom",
    "yōkai": "Jókai",
}

PLACE_SEEDS = {
    "africa": ("import-region-africa", "Africa", 20.0, 0.0, "Region", "Regional"),
    "america": ("import-region-americas", "Americas", -75.0, 10.0, "Region", "Regional"),
    "americas": ("import-region-americas", "Americas", -75.0, 10.0, "Region", "Regional"),
    "anatolia": ("import-region-anatolia", "Anatolia", 35.0, 39.0, "Region", "Regional"),
    "andes": ("import-region-andes", "Andes", -72.0, -13.0, "Region", "Regional"),
    "antarctica": ("import-region-antarctica", "Antarctica", 0.0, -82.0, "Continent", "Regional"),
    "arabia": ("import-region-arabia", "Arabia", 45.0, 24.0, "Region", "Regional"),
    "asia": ("import-region-asia", "Asia", 100.0, 34.0, "Continent", "Regional"),
    "atlantic": ("import-region-atlantic-ocean", "Atlantic Ocean", -30.0, 25.0, "Ocean", "Regional"),
    "australia": ("import-region-australia", "Australia", 134.0, -25.0, "Region", "Regional"),
    "britain": ("import-region-britain", "Britain", -2.0, 54.0, "Region", "Regional"),
    "celtic": ("import-region-celtic-europe", "Celtic Europe", -4.0, 53.0, "Region", "Mythic"),
    "central-and-western-europe": (
        "import-region-central-and-western-europe",
        "Central and Western Europe",
        7.0,
        48.0,
        "Region",
        "Regional",
    ),
    "china": ("import-region-china", "China", 104.0, 35.0, "Country", "Regional"),
    "chinese": ("import-region-china", "China", 104.0, 35.0, "Country", "Mythic"),
    "eastern-mediterranean": (
        "import-region-eastern-mediterranean",
        "Eastern Mediterranean",
        32.0,
        35.0,
        "Region",
        "Regional",
    ),
    "egypt": ("import-region-egypt", "Egypt", 30.0, 26.0, "Country", "Regional"),
    "egyptian": ("import-region-egypt", "Egypt", 30.0, 26.0, "Country", "Mythic"),
    "england": ("import-region-england", "England", -1.5, 52.4, "Country", "Regional"),
    "eurasia": ("import-region-eurasia", "Eurasia", 60.0, 50.0, "Region", "Regional"),
    "europe": ("import-region-europe", "Europe", 10.0, 50.0, "Continent", "Regional"),
    "france": ("import-region-france", "France", 2.0, 46.0, "Country", "Regional"),
    "greek": ("import-region-greece", "Greece", 22.0, 39.0, "Country", "Mythic"),
    "greece": ("import-region-greece", "Greece", 22.0, 39.0, "Country", "Regional"),
    "iceland": ("import-region-iceland", "Iceland", -19.0, 65.0, "Country", "Regional"),
    "india": ("import-region-india", "India", 78.0, 22.0, "Country", "Regional"),
    "ireland": ("import-region-ireland", "Ireland", -8.0, 53.0, "Country", "Regional"),
    "italy": ("import-region-italy", "Italy", 12.5, 42.5, "Country", "Regional"),
    "japan": ("import-region-japan", "Japan", 138.0, 37.0, "Country", "Regional"),
    "japanese": ("import-region-japan", "Japan", 138.0, 37.0, "Country", "Mythic"),
    "levant": ("import-region-levant", "Levant", 36.0, 33.0, "Region", "Regional"),
    "mediterranean": ("import-region-mediterranean", "Mediterranean", 18.0, 37.0, "Region", "Regional"),
    "mesopotamia": ("import-region-mesopotamia", "Mesopotamia", 44.0, 33.0, "Region", "Regional"),
    "mexico": ("import-region-mexico", "Mexico", -102.0, 23.0, "Country", "Regional"),
    "middle-east": ("import-region-middle-east", "Middle East", 45.0, 29.0, "Region", "Regional"),
    "north-africa": ("import-region-north-africa", "North Africa", 10.0, 28.0, "Region", "Regional"),
    "north-america": ("import-region-north-america", "North America", -100.0, 45.0, "Continent", "Regional"),
    "norse": ("import-region-northern-europe", "Northern Europe", 15.0, 60.0, "Region", "Mythic"),
    "northern-europe": ("import-region-northern-europe", "Northern Europe", 15.0, 60.0, "Region", "Regional"),
    "pacific": ("import-region-pacific-ocean", "Pacific Ocean", -150.0, 0.0, "Ocean", "Regional"),
    "persia": ("import-region-persia", "Persia", 53.0, 32.0, "Region", "Regional"),
    "roman-empire": ("import-region-roman-empire", "Roman Empire", 12.5, 42.0, "Region", "Regional"),
    "roman-judea": ("import-region-roman-judea", "Roman Judea", 35.0, 32.0, "Region", "Regional"),
    "rome": ("import-region-rome", "Rome", 12.496, 41.902, "City", "Regional"),
    "russia": ("import-region-russia", "Russia", 90.0, 60.0, "Country", "Regional"),
    "slavic": ("import-region-slavic-europe", "Slavic Europe", 25.0, 52.0, "Region", "Mythic"),
    "slavic-europe": ("import-region-slavic-europe", "Slavic Europe", 25.0, 52.0, "Region", "Regional"),
    "south-asia": ("import-region-south-asia", "South Asia", 78.0, 22.0, "Region", "Regional"),
    "southwest-asia": ("import-region-southwest-asia", "Southwest Asia", 45.0, 29.0, "Region", "Regional"),
    "ukraine": ("import-region-ukraine", "Ukraine", 31.0, 49.0, "Country", "Regional"),
    "united-states": ("import-region-united-states", "United States", -98.0, 39.0, "Country", "Regional"),
    "wales": ("import-region-wales", "Wales", -3.8, 52.3, "Country", "Regional"),
    "western-europe": ("import-region-western-europe", "Western Europe", 2.0, 48.0, "Region", "Regional"),
}

SKIPPED_PLACE_LABELS = {
    "global",
    "global-oceans",
    "moon",
    "space",
    "moon-space",
    "science",
    "technology",
    "present-day",
}


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    chars: list[str] = []
    for char in normalized:
        if unicodedata.category(char) == "Mn":
            continue
        chars.append(char.lower() if char.isalnum() else "-")
    return re.sub("-+", "-", "".join(chars)).strip("-") or "entry"


def translation_key(value: str) -> str:
    return " ".join(value.strip().lower().split())


def cs_label(value: str) -> str:
    return CS_LABELS.get(translation_key(value), value)


def label_translations(value: str) -> dict[str, str]:
    return {
        "en": value,
        "cs": cs_label(value),
    }


def value(row: tuple[object, ...], index: dict[str, int], header: str) -> str:
    offset = index.get(header)
    if offset is None or offset >= len(row) or row[offset] is None:
        return ""
    return str(row[offset]).strip()


def infer_kind(category: str) -> str:
    lower = category.lower()
    if "mythology" in lower:
        return "MythologyStory"
    if "invention" in lower:
        return "Invention"
    if "exploration" in lower:
        return "Exploration"
    if "war" in lower:
        return "War"
    if "civilization" in lower:
        return "Civilization"
    if "science" in lower:
        return "ScientificConcept"
    if "technology" in lower:
        return "Technology"
    return "Event"


def infer_icon_key(kind: str, *labels: str) -> str:
    text = " ".join(labels).lower()
    if "mythology" in text or kind in {"MythologyStory", "MythologyEntity"}:
        return "game-icons:greek-temple"
    if "war" in text or kind == "War":
        return "game-icons:crossed-swords"
    if "exploration" in text or kind == "Exploration":
        return "mdi:compass-outline"
    if "invention" in text or kind == "Invention":
        return "mdi:lightbulb-on-outline"
    if "technology" in text or kind == "Technology":
        return "mdi:chip"
    if "science" in text or kind == "ScientificConcept":
        return "mdi:atom"
    if "civilization" in text or kind == "Civilization":
        return "mdi:city-variant-outline"
    if "religion" in text:
        return "mdi:book-cross"
    if "space" in text:
        return "mdi:rocket-launch-outline"
    if "politics" in text:
        return "mdi:account-group-outline"
    return "mdi:timeline-clock-outline"


def tags_from_value(raw: str, group: str) -> list[dict[str, object]]:
    tags = []
    for part in raw.split("/"):
        name = part.strip()
        if not name:
            continue
        tags.append({
            "slug": slugify(f"{group}-{name}"),
            "group": group,
            "translations": label_translations(name),
        })
    return tags


def places_from_value(raw: str, source_type: str) -> list[dict[str, object]]:
    places = []
    seen = set()
    for part in raw.split("/"):
        label = part.replace("and other regions", "").replace("other regions", "").strip()
        lookup = slugify(label)
        if not lookup or lookup in SKIPPED_PLACE_LABELS or lookup in seen or lookup not in PLACE_SEEDS:
            continue
        seen.add(lookup)
        seed_slug, name, longitude, latitude, place_type, confidence = PLACE_SEEDS[lookup]
        places.append(
            {
                "slug": seed_slug,
                "translations": label_translations(name),
                "role": "Region",
                "placeType": place_type,
                "spatialConfidence": confidence,
                "longitude": longitude,
                "latitude": latitude,
                "note": f"Approximate {source_type} centroid inferred from workbook value '{raw}'.",
            }
        )
    return places


def audio_for_slug(slug: str, audio_root: Path) -> list[dict[str, object]]:
    tracks = []
    if not audio_root.exists():
        return tracks
    for lang_dir in sorted(path for path in audio_root.iterdir() if path.is_dir()):
        language = LANGUAGE_ALIASES.get(lang_dir.name.lower())
        if language is None:
            continue
        mp3_root = lang_dir / "mp3"
        for candidate in (mp3_root / f"{slug}.{language}.mp3", mp3_root / f"{slug}.mp3"):
            if candidate.exists() and candidate.stat().st_size > 0:
                tracks.append(
                    {
                        "languageCode": language,
                        "kind": "Narration",
                        "isPrimary": True,
                        "sortOrder": 0,
                        "title": f"{slug} narration",
                        "path": f"audio/{language}/{slug}.mp3",
                        "_sourcePath": str(candidate),
                    }
                )
                break
    return tracks


def images_for_slug(slug: str, images_root: Path) -> list[dict[str, object]]:
    if not images_root.exists():
        return []
    matches = sorted(
        path
        for path in images_root.iterdir()
        if path.is_file() and path.stem.lower() == slug.lower() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif"}
    )
    images = []
    for index, image_path in enumerate(matches):
        images.append(
            {
                "kind": "Primary" if index == 0 else "Gallery",
                "isPrimary": index == 0,
                "sortOrder": index,
                "languageCode": "en",
                "altText": slug.replace("-", " "),
                "path": f"images/{image_path.name}",
                "_sourcePath": str(image_path),
            }
        )
    return images


def time_period_from_era(era: str) -> dict[str, object] | None:
    if not era:
        return None

    slug = slugify(era)
    start_year, end_year = ERA_RANGES.get(slug, (None, None))
    return {
        "slug": slug,
        "translations": label_translations(era),
        "periodType": "Era",
        "relationType": "Primary",
        "startYear": start_year,
        "endYear": end_year,
    }


def read_entries(workbook_path: Path, audio_root: Path, images_root: Path) -> dict[str, list[dict[str, object]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    package_entries = {package_slug: [] for package_slug in SUPPORTED_SHEETS.values()}

    for sheet_name, package_slug in SUPPORTED_SHEETS.items():
        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(iterator)]
        index = {header: offset for offset, header in enumerate(headers) if header}
        for row_number, row in enumerate(iterator, start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            if sheet_name == "Master Timeline":
                title = value(row, index, "Event / development") or f"{sheet_name} row {row_number}"
                slug = slugify(title)
                category = value(row, index, "Category")
                region = value(row, index, "Region")
                kind = infer_kind(category)
                era = value(row, index, "Era")
                date_label = value(row, index, "Approx. date")
                dating_note = value(row, index, "Dating confidence")
                summary = value(row, index, "Why it matters")
                source_url = value(row, index, "Source URL")
                entry = {
                    "slug": slug,
                    "sourceSheet": sheet_name,
                    "sourceRow": row_number,
                    "kind": kind,
                    "iconKey": infer_icon_key(kind, category, region),
                    "status": "Published",
                    "realityStatus": "Mythological" if "mythology" in category.lower() else "Historical",
                    "title": title,
                    "dateLabel": date_label,
                    "timeConfidence": dating_note,
                    "translations": {
                        "en": {
                            "title": title,
                            "summary": summary,
                            "whyItMatters": summary,
                            "datingNote": dating_note,
                        }
                    },
                    "tags": tags_from_value(category, "category") + tags_from_value(region, "legacy-region-label"),
                    "timePeriods": [time_period_from_era(era)] if time_period_from_era(era) else [],
                    "sources": [{"url": source_url, "supportsField": "General"}] if source_url else [],
                    "places": places_from_value(region, "region"),
                    "raw": {header: value(row, index, header) for header in index},
                }
            else:
                title = value(row, index, "Figure / creature") or f"{sheet_name} row {row_number}"
                slug = slugify(title)
                tradition = value(row, index, "Tradition")
                mythology_type = value(row, index, "Type")
                date_label = value(row, index, "Probable tradition age")
                dating_note = value(row, index, "Dating note")
                source_url = value(row, index, "Source URL")
                entry = {
                    "slug": slug,
                    "sourceSheet": sheet_name,
                    "sourceRow": row_number,
                    "kind": "MythologyEntity",
                    "iconKey": infer_icon_key("MythologyEntity", tradition, mythology_type),
                    "status": "Published",
                    "realityStatus": "Mythological",
                    "title": title,
                    "dateLabel": date_label,
                    "timeConfidence": dating_note,
                    "translations": {
                        "en": {
                            "title": title,
                            "summary": value(row, index, "What it represents / famous story"),
                            "description": value(row, index, "Earliest important written evidence"),
                            "datingNote": dating_note,
                        }
                    },
                    "tags": [{"slug": "category-mythology", "group": "category", "translations": label_translations("Mythology")}]
                    + tags_from_value(tradition, "tradition")
                    + tags_from_value(mythology_type, "mythology-type"),
                    "timePeriods": [],
                    "sources": [{"url": source_url, "supportsField": "General"}] if source_url else [],
                    "places": places_from_value(tradition, "tradition"),
                    "raw": {header: value(row, index, header) for header in index},
                }

            entry["audio"] = [{key: val for key, val in track.items() if key != "_sourcePath"} for track in audio_for_slug(slug, audio_root)]
            entry["images"] = [{key: val for key, val in image.items() if key != "_sourcePath"} for image in images_for_slug(slug, images_root)]
            entry["_media"] = {
                "audio": audio_for_slug(slug, audio_root),
                "images": images_for_slug(slug, images_root),
            }
            package_entries[package_slug].append(entry)

    return package_entries


def write_package(package_slug: str, entries: list[dict[str, object]], out_root: Path) -> Path:
    package_dir = out_root / package_slug
    if package_dir.exists():
        shutil.rmtree(package_dir)
    package_dir.mkdir(parents=True)

    document_entries = []
    media_files: list[tuple[Path, str]] = []
    for entry in entries:
        document_entry = {key: value for key, value in entry.items() if key != "_media"}
        document_entries.append(document_entry)
        media = entry.get("_media", {})
        for track in media.get("audio", []):
            media_files.append((Path(track["_sourcePath"]), track["path"]))
        for image in media.get("images", []):
            media_files.append((Path(image["_sourcePath"]), image["path"]))

    document = {
        "schemaVersion": 1,
        "packageSlug": package_slug,
        "title": "Master Timeline" if package_slug == "master-timeline" else "Mythology",
        "defaultLanguage": "en",
        "entries": document_entries,
    }
    (package_dir / "entries.json").write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")

    for source_path, relative_path in media_files:
        target_path = package_dir / relative_path
        target_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, target_path)

    zip_path = out_root / f"{package_slug}.zip"
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in package_dir.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(package_dir).as_posix())

    return zip_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Build content package ZIPs from the workbook and generated media.")
    parser.add_argument("--workbook", default="world_history_mythology_timeline_to_2026.xlsx")
    parser.add_argument("--audio-root", default="generated/audio")
    parser.add_argument("--images-root", default="generated/images")
    parser.add_argument("--out", default="generated/packages")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    audio_root = Path(args.audio_root)
    images_root = Path(args.images_root)
    out_root = Path(args.out)
    out_root.mkdir(parents=True, exist_ok=True)

    packages = read_entries(workbook_path, audio_root, images_root)
    for package_slug, entries in packages.items():
        zip_path = write_package(package_slug, entries, out_root)
        audio_count = sum(len(entry.get("audio", [])) for entry in entries)
        image_count = sum(len(entry.get("images", [])) for entry in entries)
        print(f"Wrote {zip_path} with {len(entries)} entries, {audio_count} audio files, {image_count} images.")


if __name__ == "__main__":
    main()
